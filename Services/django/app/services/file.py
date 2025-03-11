import os

import boto3
from docx import Document
from openpyxl import load_workbook

from app.models import Message
from django.conf import settings
from django.contrib.auth.models import User

BUCKET_NAME = "test"
s3 = boto3.client(
	"s3",
	endpoint_url=settings.AWS_S3_ENDPOINT_URL,
	aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
	aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
)


def create_bucket(bucket_name: str):
	s3.create_bucket(
		Bucket=bucket_name,
		CreateBucketConfiguration={"LocationConstraint": "us-east-1"},
	)


class FileService:
	@staticmethod
	def create_message(sender: User, room_id: int, file):
		message = Message.objects.create(sender=sender, room_id=room_id, file=file)

		file_name = message.file.path
		key_name = message.file.path

		s3.upload_file(file_name, BUCKET_NAME, key_name)
		os.remove(message.file.path)
		return message

	@staticmethod
	def get_file(message_id):
		message = Message.objects.filter(pk=message_id).first()

		s3.download_file(
			BUCKET_NAME,
			message.file.path,
			message.file.path,
		)

		if message.file.path.endswith(".docx"):
			doc = Document(message.file.path)
			content = []
			for paragraph in doc.paragraphs:
				content.append(paragraph.text)
		elif message.file.path.endswith(".xlsx"):
			workbook = load_workbook(message.file.path)
			sheet = workbook.active
			content = []
			for row in sheet.iter_rows(values_only=True):
				content.append(row)
		else:
			with open(message.file.path, encoding="utf-8", errors="ignore") as f:
				content = f.readlines()

		return {"name": os.path.basename(message.file.path), "content": content}
